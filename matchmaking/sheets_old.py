from openpyxl import Workbook, load_workbook
from matchmaking import match, player
from settings import *


def find_player(discord_user, path=playerbase):
    player_id = str(discord_user.id)
    wb = load_workbook(path)
    ws = wb.active
    player_row = None

    for row in range(1, ws.max_row+1):
        if player_id == ws['S' + str(row)].value:
            player_row = row

    class_fix = {
        'Cavalry': 'cav',
        'Archer': 'arch',
        'Infantry': 'inf'
    }
    if player_row:
        stats = {
            'discord_id': player_id,
            'mmr': int(ws['R' + str(player_row)].value),
            'main': class_fix[ws['D' + str(player_row)].value],
            'secondary': class_fix[ws['E' + str(player_row)].value],
            'nickname': ws['B' + str(player_row)].value
        }

        founded_player = player.Player(stats)
        return founded_player
    else:
        return None


def add_player(discord_member, path=playerbase):
    wb = load_workbook(path)
    ws = wb.active
    ws.append([(str(len(ws['R']))), discord_member.name, '', 'Infantry', 'Archer', '0', '0',
              '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '1000', str(discord_member.id)])
    wb.save(playerbase)
